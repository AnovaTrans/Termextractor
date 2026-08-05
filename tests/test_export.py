"""Every export format must produce valid output.

Regressions fixed here (all three failed while only JSON worked):
- XLSX: Font used in a staticmethod but imported only in another method.
- CSV: csv.DictWriter given an `encoding` kwarg it does not accept, writing to
  a BytesIO it cannot write to.
- TBX: set('xml:lang', ...) rejected — needs the expanded namespace form.
Run with: python -m pytest tests/test_export.py
"""
import io
import xml.etree.ElementTree as ET

import pytest
from openpyxl import load_workbook

from src.models import (Term, ExtractionResult, DerivativeStatistics,
                        BilinguialLookupStatistics)
from src.io import FormatExporter


@pytest.fixture
def result():
    terms = [
        Term(term="River West OPEN", translation="Ривър Уест", domain="Real Estate",
             subdomain="Retail", pos="NOUN", definition="a retail park",
             context="the River West OPEN park", relevance_score=95,
             confidence_score=90, frequency=3),
        Term(term="Control Panels", translation="Контролни панели", domain="Engineering",
             pos="NOUN", definition="custom-built", context="Control Panels by Metta",
             relevance_score=88, confidence_score=85, frequency=1),
    ]
    res = ExtractionResult(terms=terms, domain_hierarchy=["Real Estate"],
                           source_language="en", target_language="bg")
    res.statistics = res._calculate_statistics(terms)
    # As set by the real extraction flow: these carry list-valued fields
    # (e.g. modes_used=[]) that openpyxl cannot put in a cell directly.
    res.lookup_statistics = BilinguialLookupStatistics().__dict__
    res.derivative_statistics = DerivativeStatistics().__dict__
    return res


def test_xlsx_is_a_valid_workbook_with_a_bold_header(result):
    data = FormatExporter().export(result, "xlsx")
    assert data[:4] == b"PK\x03\x04"
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Terms"]
    assert ws.cell(1, 1).value == "Term"
    assert ws.cell(1, 1).font.bold is True          # the Font that used to be undefined
    assert ws.cell(2, 1).value == "River West OPEN"


def test_csv_has_a_bom_headers_and_unicode(result):
    data = FormatExporter().export(result, "csv")
    assert data.startswith(b"\xef\xbb\xbf")          # utf-8-sig BOM for Excel
    text = data.decode("utf-8-sig")
    assert text.splitlines()[0].startswith("term,")
    assert "Ривър Уест" in text                       # unicode survived


def test_xlsx_handles_list_valued_statistics(result):
    # modes_used=[] and other list fields must not raise "Cannot convert [] to Excel".
    data = FormatExporter().export(result, "xlsx")
    wb = load_workbook(io.BytesIO(data))
    assert "Statistics" in wb.sheetnames
    # every populated cell is a scalar openpyxl accepted
    for row in wb["Statistics"].iter_rows():
        for cell in row:
            assert not isinstance(cell.value, (list, tuple, dict, set))


def test_csv_of_no_terms_does_not_crash():
    empty = ExtractionResult(terms=[])
    data = FormatExporter().export(empty, "csv")
    assert b"No terms" in data


def test_tbx_is_well_formed_and_uses_xml_lang(result):
    data = FormatExporter().export(result, "tbx")
    root = ET.fromstring(data)                        # raises if malformed
    xml_lang = "{http://www.w3.org/XML/1998/namespace}lang"
    assert root.get(xml_lang) == "en"
    lang_sets = root.iter("langSet")
    langs = {ls.get(xml_lang) for ls in lang_sets}
    assert "en" in langs and "bg" in langs            # source + target


def test_json_round_trips(result):
    import json
    data = json.loads(FormatExporter().export(result, "json").decode("utf-8"))
    assert data["metadata"]["source_language"] == "en"
    assert len(data["terms"]) == 2


def test_every_format_returns_nonempty_bytes(result):
    ex = FormatExporter()
    for fmt in ("xlsx", "csv", "tbx", "json"):
        data = ex.export(result, fmt)
        assert isinstance(data, bytes) and len(data) > 0
